import frappe, erpnext, json
from frappe.utils import (
	cstr, get_link_to_form
)
from frappe import _
import openpyxl as xl
import time
import datetime
import calendar
from datetime import timedelta
from copy import copy
from pathlib import Path
from itertools import groupby
from operator import itemgetter
from one_fm.processor import sendemail
from one_fm.utils import production_domain

def get_basic_salary(employee):
	filters = {
		'docstatus': 1,
		'employee': employee
	}
	salary_structure = frappe.get_value("Salary Structure Assignment", filters, "salary_structure", order_by="from_date desc")
	if salary_structure:
		basic_salary = frappe.db.sql("""
			SELECT amount FROM `tabSalary Detail`
			WHERE parenttype="Salary Structure"
			AND parent=%s
			AND salary_component="Basic"
		""",(salary_structure), as_dict=1)

		return basic_salary[0].amount if len(basic_salary) > 0 else 0.00
	else:
		frappe.throw(_("No Assigned Salary Structure found for the selected employee."))

@frappe.whitelist()
def export_payroll(doc, method):
	""" This method fetches the company bank details and makes a call to the export function based on the provided bank.

	Args:
		payroll_entry (document object): The payroll entry document object to set the export file field for.
	"""
	# Check if Export is enabled.
	if frappe.db.get_single_value("HR and Payroll Additional Settings", 'enable_export'):
		# Get default bank used to pay salaries
		default_bank = frappe.db.get_single_value("HR and Payroll Additional Settings", 'default_bank')

		# Fetch template and bank code for default bank
		template_path, bank_code = frappe.db.get_value("Bank", {'name': default_bank}, ["payroll_export_template", "bank_code"])

		cash_salary_employees = []

		for employee in doc.employees:
			if employee.salary_mode == "Cash":
				cash_salary_employees.append(employee)
			elif employee.salary_mode == "Bank":
				if not employee.iban_number:
					frappe.throw(_("No Iban/Bank account set for employee: {employee}".format(employee=employee.employee)))
			elif not employee.salary_mode:
				frappe.throw(_("No salary mode set for employee: {employee}".format(employee=employee.employee)))

		if "NBK" in bank_code:
			# Enqueue method for longer list of employees
			if len(doc.employees) > 30:
				frappe.enqueue(export_nbk, doc=doc, template_path=template_path)
			else:
				export_nbk(doc, template_path)

		if len(cash_salary_employees) > 0:
			if len(cash_salary_employees) > 30:
				frappe.enqueue(export_cash_payroll, cash_salary_employees=cash_salary_employees, doc_name=doc.name)
			else:
				export_cash_payroll(cash_salary_employees, doc.name)
		else:
			frappe.msgprint(_("No employees with salary mode as Cash."))

def export_nbk(doc, template_path):
	"""This method fetches the bank template from the provided directory, copies the template style and data into a new workbook, writes payroll entry data
		into the new workbook and saves it in the public files directory of the current site.

	Args:
		payroll_entry (document object): The payroll entry document object to be used for exporting the payroll data into the provided bank template and set the export file field.
		template_path (str): Path to the bank template file
	"""

	start = time.time()

	employees = doc.employees

	if len(employees) == 0:
		frappe.throw(_("No employees added to payroll entry."))

	if not doc.bank_account:
		frappe.throw(_("No bank account set in payroll entry."))

	iban, bank_account_no = frappe.db.get_value("Bank Account", {'name': doc.bank_account}, ["iban", "bank_account_no"])

	if not iban and not bank_account_no:
		frappe.throw(_("No IBAN or bank account number set for Bank Account: {bank_account}".format(bank_account=doc.bank_account)))

	try:
		# Load template file
		template_filename = cstr(frappe.local.site) + template_path
		template_wb = xl.load_workbook(filename=template_filename)
		template_ws = template_wb.worksheets[0]

		#-------------------- Copy template data to destination worksheet --------------------#
		# Setup new file
		destination_wb = xl.Workbook()
		destination_ws = destination_wb.active

		# Max row number with template data as per NBK template
		mr = 12
		# Max column number with template data as per NBK template
		mc = 7

		# copying the cell values from source excel file to destination excel file
		for i in range (1, mr + 1):
			for j in range (1, mc + 1):
				# reading cell value from source excel file
				c = template_ws.cell(row = i, column = j)

				d = destination_ws.cell(row = i, column = j)
				# writing the read value to destination excel file
				d.value = c.value
				# Copy cell style
				if c.has_style:
					d.font = copy(c.font)
					d.border = copy(c.border)
					d.fill = copy(c.fill)
					d.number_format = copy(c.number_format)
					d.protection = copy(c.protection)
					d.alignment = copy(c.alignment)
		#---------------------- End copy template data to destination worksheet ------------------#

		# Currency map as per NBK bank template
		currency_map = {
			'KWD': 'KWD - Kuwaiti Dinar',
			'USD': 'USD - US Dollar',
			'GBP': 'GBP - British Pound',
			'EUR': 'EUR - EURO',
			'CAD': 'CAD - Canadian Dollar',
			'AUD': 'AUD - Australian Dollar'
		}

  		# Set column numbers based on NBK bank template
		source_ws_emp_column_map = {
			'Employee Number': 1,
			'Employee Name': 2,
			'Employee IBAN Number': 3,
			'Payment Amount': 4,
			'Bank Code': 5,
			'Employee Civil ID': 6,
			'MOSAL ID': 7
		}

		# TODO: Fetch firm number
		currency = currency_map[doc.currency]
		posting_date = cstr(doc.posting_date).split("-") # => [yyyy, mm, dd]
		payment_month = posting_date[1] + "-" + posting_date[0] # => mm-yyyy

		# Set basic payroll details in row and columns based on NBK bank template
		destination_ws.cell(row=3, column=3).value = doc.company
		destination_ws.cell(row=4, column=3).value = bank_account_no or iban[-10:0]
		destination_ws.cell(row=5, column=3).value = doc.payment_purpose
		destination_ws.cell(row=6, column=3).value = payment_month
		destination_ws.cell(row=7, column=3).value = currency

		# Row number to start entering employee payroll data
		source_ws_employee_row_number = 13

		# Employee count for employee number column
		employee_number_column_count = 1

		total_hash = 0
		total_amount = 0

		# Set employee payroll details
		for employee in employees:
			if employee.salary_mode == "Bank":
				destination_ws.cell(row=source_ws_employee_row_number, column=source_ws_emp_column_map["Employee Number"]).value = employee_number_column_count
				destination_ws.cell(row=source_ws_employee_row_number, column=source_ws_emp_column_map["Employee Name"]).value = employee.employee_name
				destination_ws.cell(row=source_ws_employee_row_number, column=source_ws_emp_column_map["Employee IBAN Number"]).value = employee.iban_number
				destination_ws.cell(row=source_ws_employee_row_number, column=source_ws_emp_column_map["Payment Amount"]).value = employee.payment_amount
				destination_ws.cell(row=source_ws_employee_row_number, column=source_ws_emp_column_map["Bank Code"]).value = employee.bank_code
				destination_ws.cell(row=source_ws_employee_row_number, column=source_ws_emp_column_map["Employee Civil ID"]).value = employee.civil_id_number
				destination_ws.cell(row=source_ws_employee_row_number, column=source_ws_emp_column_map["MOSAL ID"]).value = employee.mosal_id

				iban_multipier = int(employee.iban_number[-10:])
				total_hash += round(iban_multipier * employee.payment_amount, 3)
				total_amount += round(employee.payment_amount, 3)

				employee_number_column_count += 1
				source_ws_employee_row_number += 1

		destination_ws.cell(row=8, column=3).value = len(employees)
		destination_ws.cell(row=9, column=3).value = total_amount
		destination_ws.cell(row=10, column=3).value = total_hash

		# Setup destination file directory with payroll entry name as filename
		Path("/home/frappe/frappe-bench/sites/{0}/public/files/payroll-entry/".format(frappe.local.site)).mkdir(parents=True, exist_ok=True)
		destination_file = cstr(frappe.local.site) + "/public/files/payroll-entry/{payroll_entry}.xlsx".format(payroll_entry=doc.name)

		# Save updated template in same source directory
		destination_wb.save(filename=destination_file)

		end = time.time()

		# print("Total Execution Time: ", end-start)

	except Exception as e:
		frappe.throw(e)

frappe.whitelist()
def export_cash_payroll(cash_salary_employees, doc_name):
	"""This method takes the list of employees who have salary mode set as Cash and exports the payroll employee details into an excel sheet.

	Args:
		cash_salary_employees (List[dict]): payroll empployee details
		doc_name (str): Name of the payroll entry document.
	"""
	try:
        # Setup destination file directory with payroll entry name as filename
		Path("/home/frappe/frappe-bench/sites/{0}/public/files/payroll-entry/".format(frappe.local.site)).mkdir(parents=True, exist_ok=True)
		destination_file = cstr(frappe.local.site) + "/public/files/payroll-entry/Cash-{payroll_entry}.xlsx".format(payroll_entry=doc_name)
		destination_wb = xl.Workbook()
		destination_ws = destination_wb.active

		# Fill color in first row
		color_fill = xl.styles.PatternFill(start_color='FFFF00',
                   end_color='FFFF00',
                   fill_type='solid')
		i = 1
		while(i < 5):
			destination_ws.cell(row=1, column=i).fill = color_fill
			i += 1

		# Set column names
		destination_ws.cell(row=1, column=1).value = "Employee Name"
		destination_ws.cell(row=1, column=2).value = "Payment Amount"
		destination_ws.cell(row=1, column=3).value = "Civil ID"
		destination_ws.cell(row=1, column=4).value = "Mosal ID"

		row_number = 2

		# Fill employees in rows
		for employee in cash_salary_employees:
			destination_ws.cell(row=row_number, column=1).value = employee.employee_name
			destination_ws.cell(row=row_number, column=2).value = employee.payment_amount
			destination_ws.cell(row=row_number, column=3).value = employee.civil_id_number
			destination_ws.cell(row=row_number, column=4).value = employee.mosal_id

			row_number += 1

		destination_wb.save(filename=destination_file)

	except Exception as e:
		frappe.log_error(e)

@frappe.whitelist()
def email_missing_payment_information(recipients):
	"""
		Send missing salary payment information
		as an email.
	"""
	# print(frappe.session.data)
	# print(recipients, '\n\n\n')

def log_payroll_failure(process, payroll_entry, error):
	error_log = frappe.log_error(
		title=_("Salary Slip {0} failed for Payroll Entry {1}").format(process, payroll_entry.name)
	)
	message_log = frappe.message_log.pop() if frappe.message_log else str(error)

	try:
		error_message = json.loads(message_log).get("message")
	except Exception:
		error_message = message_log

	error_message += "\n" + _("Check Error Log {0} for more details.").format(
		get_link_to_form("Error Log", error_log.name)
	)

	payroll_entry.db_set({"error_message": error_message, "status": "Failed"})

def create_salary_slips_for_employees(employees, payroll_entry, publish_progress=True ):
	from one_fm.api.doc_methods.salary_slip import get_ot_days
	try:
		payroll_entry = frappe.get_doc("Payroll Entry", payroll_entry.name)
		args = frappe._dict(
			{
				"salary_slip_based_on_timesheet": payroll_entry.salary_slip_based_on_timesheet,
				"payroll_frequency": payroll_entry.payroll_frequency,
				"company": payroll_entry.company,
				"posting_date": payroll_entry.posting_date,
				"deduct_tax_for_unclaimed_employee_benefits": payroll_entry.deduct_tax_for_unclaimed_employee_benefits,
				"deduct_tax_for_unsubmitted_tax_exemption_proof": payroll_entry.deduct_tax_for_unsubmitted_tax_exemption_proof,
				"exchange_rate": payroll_entry.exchange_rate,
				"currency": payroll_entry.currency,
				"payroll_entry": payroll_entry.name,
				"payroll_type": payroll_entry.payroll_type
			}
		)
		salary_slips_exist_for = get_existing_salary_slips(employees, args)
		count = 0
		start_date = payroll_entry.start_date
		end_date = payroll_entry.end_date
		salary_slip_chunk = []
		chunk_counter = 0

		employees_list = seperate_salary_slip(employees, start_date, end_date)
		if payroll_entry.payroll_type == "Over-Time":
			for emp in employees_list:
				if get_ot_days(emp['employee'], emp['start_date'], emp['end_date']) == 0:
					employees_list.remove(emp)

		if len(employees_list) < 30:
			for emp in employees_list:
				if emp['employee'] not in salary_slips_exist_for:
					args.update({"doctype": "Salary Slip"})
					args.update(emp)
					frappe.get_doc(args).insert()
					count += 1
					if publish_progress:
						frappe.publish_progress(
							count * 100 / len(set(employees) - set(salary_slips_exist_for)),
							title=_("Creating Salary Slips..."),
						)

		else:
			for emp in employees_list:
				if emp['employee'] not in salary_slips_exist_for:
					args.update({"doctype": "Salary Slip"})
					args.update(emp)

					# salary_slip_list.append(frappe.get_doc(args))
					salary_slip_chunk.append(frappe.get_doc(args))
					chunk_counter += 1
					if len(salary_slip_chunk) >= 30:
						frappe.enqueue(create_salary_slip_chunk,slips=salary_slip_chunk.copy(), queue="long")
						salary_slip_chunk = []
						chunk_counter=0

					# frappe.get_doc(args).insert()

			if salary_slip_chunk:
				frappe.enqueue(create_salary_slip_chunk,slips=salary_slip_chunk, queue="long")
		payroll_entry.db_set({"status": "Submitted", "salary_slips_created": 1, "error_message": ""})

		if salary_slips_exist_for:
			frappe.msgprint(
				_(
					"Salary Slips already exist for employees {}, and will not be processed by this payroll."
				).format(frappe.bold(", ".join(emp for emp in salary_slips_exist_for))),
				title=_("Message"),
				indicator="orange",
			)

	except Exception as e:
		frappe.db.rollback()
		log_payroll_failure("creation", payroll_entry, e)

	finally:
		frappe.db.commit()  # nosemgrep
		frappe.publish_realtime("completed_salary_slip_creation")

def create_salary_slip_chunk(slips):
	for slip in slips:
		slip.insert()
		slip.save()

@frappe.whitelist()
def check_salary_slip_count(doc):
	payroll_entry = frappe.get_doc("Payroll Entry", doc)
	if payroll_entry.salary_slips_created == 1:
		salary_count = frappe.db.sql_list(
			f"""
			select Count(distinct employee) as salary_count from `tabSalary Slip`
			where docstatus!= 2
				and company = '{payroll_entry.company}'
				and start_date = '{payroll_entry.start_date}'
				and end_date = '{payroll_entry.end_date}'
			""", as_dict=1)
		print(salary_count[0])
		if salary_count[0] != payroll_entry.number_of_employees:

			payroll_entry.db_set({"status": "Pending Salary Slip", "error_message": ""})
		else:
			payroll_entry.db_set({"status": "Submitted", "error_message": ""})
		frappe.db.commit()
	else:
		payroll_entry.db_set({"status": "Draft", "error_message": ""})
	return True

@frappe.whitelist()
def create_pending_sal_slip(doc):
	payroll_entry = frappe.get_doc("Payroll Entry", doc)
	pe_employees = [emp.employee for emp in payroll_entry.employees]
	ss_employees = frappe.db.sql_list(
		f"""
		select distinct employee from `tabSalary Slip`
		where docstatus!= 2
			and company = '{payroll_entry.company}'
			and start_date >= '{payroll_entry.start_date}'
			and end_date <='{payroll_entry.end_date}'
		""", as_dict=True)
	employees = [e for e in pe_employees if e not in ss_employees]

	if len(employees) != 0:
		if len(employees) > 30 or frappe.flags.enqueue_payroll_entry:
			payroll_entry.db_set("status", "Queued")
			frappe.enqueue(
				create_salary_slips_for_employees,
				employees=employees,
				payroll_entry=payroll_entry,
				publish_progress=False,
				timeout=6000,
				queue='long'
			)
			frappe.msgprint(
				_("Salary Slip creation is queued. It may take a few minutes"),
				alert=True,
				indicator="blue",
			)
		else:
			create_salary_slips_for_employees(employees, payroll_entry=payroll_entry, publish_progress=False)
			# since this method is called via frm.call this doc needs to be updated manually
			payroll_entry.reload()
	return True


def get_existing_salary_slips(employees, args):
	return frappe.db.sql_list(
		"""
		select distinct employee from `tabSalary Slip`
		where docstatus!= 2 and company = %s and payroll_entry = %s
			and start_date >= %s and end_date <= %s
			and employee in (%s)
	"""
		% ("%s", "%s", "%s", "%s", ", ".join(["%s"] * len(employees))),
		[args.company, args.payroll_entry, args.start_date, args.end_date] + employees,
	)

def seperate_salary_slip(employees, start_date, end_date):
	parm = []
	for emp in employees:
		salary_structure_assignment = frappe.get_all("Salary Structure Assignment", {"employee":emp, "from_date":["between",(start_date, end_date)]},["*"])

		if len(salary_structure_assignment) > 1:
			mid_date = ""
			for ssa in salary_structure_assignment:
				start_date = frappe.utils.get_datetime(start_date).date()
				end_date = frappe.utils.get_datetime(end_date).date()

				if ssa.from_date > start_date and ssa.from_date < end_date:
					mid_date = ssa.from_date
			if mid_date:
				parm.append({"employee": emp, "start_date":start_date, "end_date":mid_date - timedelta(days=1)})
				parm.append({"employee": emp, "start_date":mid_date , "end_date":end_date})
		else:
			parm.append({"employee": emp, "start_date":start_date , "end_date":end_date})

	return parm

def notify_for_open_leave_application():
	try:
		if not frappe.db.get_single_value('HR and Payroll Additional Settings', 'remind_open_leave_application') and not production_domain():
			return

		open_leave_application = {}
		leave_list = frappe.get_all("Leave Application", {"workflow_state":"Open"}, ['*'])
		# sort INFO data by 'leave_approver' key.
		leave_list = sorted(leave_list, key=itemgetter('leave_approver'))

		#group leave application by leave approver
		for key, value in groupby(leave_list, itemgetter('leave_approver')):
			open_leave_application[key] = []
			for k in value:
				open_leave_application[key].append(k.name)
		for ola in open_leave_application:
			recipient = [ola]
			message = "<p>The Following Leave Application needs to be approved. Kindly, take action before midnight. </p><ul>"
			for leave_id in open_leave_application[ola]:
				doc_url = frappe.utils.get_link_to_form("Leave Application", leave_id)
				message += "<li>"+doc_url+"</li>"
			message += "</ul>"

			sendemail(recipients= recipient , subject="Leave Application need approval", message=message)
	except Exception as error:
		frappe.log_error(str(error), 'Open Leave Application reminder failed')
